#!/usr/bin/env python3
"""Create one or more AESG workbook sheets from a compact JSON specification."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


GREEN = "008C95"
GRAY = "343741"
WHITE = "FFFFFF"
PURPLE = "6D2077"
RED = "DA291C"
YELLOW = "FFC72C"
FONT = "Verdana"


def skill_root() -> Path:
    return Path(__file__).resolve().parents[2]


def normalise_columns(spec: dict) -> list[dict]:
    columns = []
    for item in spec.get("columns", []):
        columns.append({"key": item, "label": item} if isinstance(item, str) else dict(item))
    if not columns:
        raise ValueError("each sheet must contain at least one column")
    return columns


def chart_for(chart_type: str):
    if chart_type == "line":
        return LineChart()
    if chart_type == "pie":
        return PieChart()
    return BarChart()


def table_name(value: str, index: int) -> str:
    clean = re.sub(r"[^A-Za-z0-9_]", "", value.replace(" ", "_"))
    if not clean or clean[0].isdigit():
        clean = f"AESG_{clean}"
    return f"{clean[:230]}_{index}"


def add_logo(worksheet) -> None:
    logo_path = skill_root() / "aesg-branding/assets/logos/aesg-brandmark-rgb.png"
    if logo_path.is_file():
        logo = Image(logo_path)
        logo.width = 130
        logo.height = 37.5
        worksheet.add_image(logo, "A1")
    worksheet.row_dimensions[1].height = 34


def row_value(row, column: dict, column_index: int, row_index: int):
    formula = column.get("formula")
    if formula:
        return str(formula).replace("{row}", str(row_index))
    if isinstance(row, dict):
        return row.get(str(column["key"]), "")
    return row[column_index - 1] if column_index - 1 < len(row) else ""


def add_chart(worksheet, spec: dict, columns: list[dict], end_row: int):
    chart_spec = spec.get("chart")
    if not chart_spec or end_row < 4:
        return len(columns), end_row
    keys = [str(column["key"]) for column in columns]
    category_key = str(chart_spec.get("category", ""))
    value_keys = chart_spec.get("values") or [chart_spec.get("value")]
    value_keys = [str(value) for value in value_keys if value is not None]
    if category_key not in keys or any(value not in keys for value in value_keys):
        raise ValueError("chart category/value must reference declared column keys")
    chart = chart_for(str(chart_spec.get("type", "bar")).casefold())
    chart.title = str(chart_spec.get("title", ""))
    chart.style = 10
    chart.height = 7
    chart.width = 12
    categories = Reference(
        worksheet, min_col=keys.index(category_key) + 1, min_row=4, max_row=end_row
    )
    for value_key in value_keys:
        value_index = keys.index(value_key) + 1
        values = Reference(worksheet, min_col=value_index, min_row=3, max_row=end_row)
        chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    series_colors = [GREEN, GRAY, PURPLE, RED, YELLOW]
    for series, colour in zip(chart.series, series_colors):
        series.graphicalProperties.solidFill = colour
        series.graphicalProperties.line.solidFill = colour
    if isinstance(chart, PieChart):
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
    worksheet.add_chart(chart, f"A{end_row + 2}")
    return max(len(columns), 8), end_row + 18


def build_sheet(workbook: Workbook, worksheet, spec: dict, sheet_index: int) -> None:
    columns = normalise_columns(spec)
    rows = list(spec.get("rows", []))
    worksheet.title = str(spec.get("sheet", f"Sheet {sheet_index}"))[:31]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A4"
    add_logo(worksheet)

    last_column = get_column_letter(len(columns))
    worksheet.merge_cells(f"A2:{last_column}2")
    title = worksheet["A2"]
    title.value = str(spec.get("title", worksheet.title))
    title.fill = PatternFill("solid", fgColor=GREEN)
    title.font = Font(name=FONT, size=14, bold=True, color=WHITE)
    title.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[2].height = 30

    thin_gray = Side(style="thin", color="D9D9D9")
    for column_index, column in enumerate(columns, start=1):
        cell = worksheet.cell(3, column_index, str(column.get("label", column["key"])))
        cell.fill = PatternFill("solid", fgColor=GRAY)
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)
        width = min(max(float(column.get("width", 18)), 8), 45)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.row_dimensions[3].height = 24.75

    for row_index, row in enumerate(rows, start=4):
        for column_index, column in enumerate(columns, start=1):
            value = row_value(row, column, column_index, row_index)
            cell = worksheet.cell(row_index, column_index, value)
            cell.font = Font(name=FONT, size=9, color=GRAY)
            cell.alignment = Alignment(
                horizontal="right" if isinstance(value, (int, float)) or str(value).startswith("=") else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=thin_gray)
            if column.get("format"):
                cell.number_format = str(column["format"])
        worksheet.row_dimensions[row_index].height = 19.5

    end_row = 3 + len(rows)
    if rows:
        table = Table(
            displayName=table_name(worksheet.title, sheet_index),
            ref=f"A3:{last_column}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    for column_index, column in enumerate(columns, start=1):
        values = column.get("validation")
        if values:
            quoted = '"' + ",".join(str(value) for value in values) + '"'
            validation = DataValidation(type="list", formula1=quoted, allow_blank=True)
            worksheet.add_data_validation(validation)
            validation.add(
                f"{get_column_letter(column_index)}4:{get_column_letter(column_index)}{max(end_row, 200)}"
            )
        if str(column.get("key", "")).casefold() == "status":
            column_letter = get_column_letter(column_index)
            worksheet.conditional_formatting.add(
                f"{column_letter}4:{column_letter}{max(end_row, 200)}",
                FormulaRule(
                    formula=[f'LOWER({column_letter}4)="at risk"'],
                    fill=PatternFill("solid", fgColor="F9DDD9"),
                    font=Font(name=FONT, color=RED),
                ),
            )

    print_last_column, print_last_row = add_chart(worksheet, spec, columns, end_row)
    worksheet.auto_filter.ref = f"A3:{last_column}{max(end_row, 3)}"
    worksheet.print_area = f"A1:{get_column_letter(print_last_column)}{max(print_last_row, 3)}"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.orientation = "landscape" if len(columns) > 6 else "portrait"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.oddFooter.center.text = "AESG | &P of &N"
    worksheet.oddFooter.center.font = f"{FONT},Regular"
    worksheet.oddFooter.center.size = 7


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--spec", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    if args.output.suffix.casefold() != ".xlsx":
        raise ValueError("--output must end in .xlsx")
    spec = json.loads(args.spec.read_text(encoding="utf-8"))
    sheet_specs = spec.get("sheets") or [spec]
    if not isinstance(sheet_specs, list) or not sheet_specs:
        raise ValueError("spec.sheets must be a non-empty list")

    workbook = Workbook()
    for index, sheet_spec in enumerate(sheet_specs, start=1):
        worksheet = workbook.active if index == 1 else workbook.create_sheet()
        combined = dict(sheet_spec)
        combined.setdefault("title", spec.get("title", combined.get("sheet", "AESG workbook")))
        build_sheet(workbook, worksheet, combined, index)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.properties.creator = "AESG"
    workbook.properties.lastModifiedBy = "AESG"
    workbook.properties.title = str(spec.get("title", "AESG workbook"))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(
        json.dumps(
            {
                "ok": True,
                "output": str(args.output.resolve()),
                "bytes": args.output.stat().st_size,
                "sheets": workbook.sheetnames,
            }
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
